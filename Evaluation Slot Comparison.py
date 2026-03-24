import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

# ── Config ───────────────────────────────────────────────────────────────────
SHEET_NAME = "Slots Allocations"
OUTPUT_FILE = "comparison_output.xlsx"


# ── Helpers ───────────────────────────────────────────────────────────────────
def extract_short_name(filename: str) -> str:
    import re
    match = re.search(r'eval_\d+', filename)
    return match.group(0) if match else os.path.splitext(filename)[0]


def make_key(row, sep="_"):
    """Concat Asset_Name + Development_Area_Name + Formation_Name + Slot_Number."""
    return sep.join(str(row[c]) for c in ["Asset_Name", "Development_Area_Name", "Formation_Name", "Slot_Number"])


def make_ltd_key(row, sep="_"):
    """Concat Asset_Name + Development_Area_Name + Formation_Name + Slot_Number."""
    return sep.join(str(row[c]) for c in ["Asset_Name", "Development_Area_Name", "Formation_Name", "Slot_Number"])


# ── Data Processing ───────────────────────────────────────────────────────────
def read_slots(file_path: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path, sheet_name=SHEET_NAME, header=0)
    except Exception as e:
        raise ValueError(f"Could not read '{SHEET_NAME}' from '{os.path.basename(file_path)}':\n{e}")

    cols = list(df.columns)
    if len(cols) < 14:
        raise ValueError(f"Not enough columns in '{os.path.basename(file_path)}' – expected at least 14, got {len(cols)}.")

    col_map = {
        cols[0]:  "Development_Area_Name",   # A
        cols[1]:  "Asset_Name",              # B
        cols[3]:  "Basin",                   # D
        cols[5]:  "Formation_Name",          # F
        cols[6]:  "Slot_Number",             # G
        cols[7]:  "slot_category",           # H
        cols[13]: "UWI",                     # N
        cols[14]: "Well_Name",               # O
        cols[19]: "Net_Acreage",             # T
        cols[24]: "Royalty_Rate",            # Y
        cols[25]: "Overriding_Royalty_Rate", # Z
    }
    df = df.rename(columns=col_map)
    df["slot_category"] = df["slot_category"].astype(str).str.strip().str.lower()
    # Remove decimals from UWI (e.g. "12345.0" → "12345")
    df["UWI"] = (pd.to_numeric(df["UWI"], errors="ignore")
                   .apply(lambda x: str(int(float(x))) if str(x).replace('.','',1).isdigit() else str(x))
                   .str.strip())
    df["source_file"]   = extract_short_name(os.path.basename(file_path))

    keep = ["source_file", "slot_category", "UWI",
            "Development_Area_Name", "Asset_Name", "Basin",
            "Formation_Name", "Slot_Number", "Well_Name",
            "Net_Acreage", "Royalty_Rate", "Overriding_Royalty_Rate"]
    return df[keep]


def read_raw(file_path: str) -> pd.DataFrame:
    """Read raw Slots Allocations sheet for appending at end."""
    return pd.read_excel(file_path, sheet_name=SHEET_NAME, header=0)


def build_summary(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]

    ltd_cols_base = ["Development_Area_Name", "Asset_Name", "Formation_Name", "Slot_Number"]

    def ltd_count(df):
        return (df[df["slot_category"] == "ltd"]
                .drop_duplicates(subset=ltd_cols_base)
                .shape[0])

    def cat_counts(df):
        grp = df.groupby("slot_category")["UWI"].nunique().reset_index()
        grp.columns = ["slot_category", "count"]
        grp = grp[grp["slot_category"] != "nan"]
        grp.loc[grp["slot_category"] == "ltd", "count"] = ltd_count(df)
        return grp.set_index("slot_category")["count"]

    c1 = cat_counts(df1)
    c2 = cat_counts(df2)
    all_cats = sorted(set(c1.index) | set(c2.index))

    rows = []
    for cat in all_cats:
        v1 = int(c1.get(cat, 0))
        v2 = int(c2.get(cat, 0))
        rows.append({
            f"Excel 1 ({file1})"  : file1,
            "Slot Category"       : cat,
            f"Count ({file1})"    : v1,
            f"Excel 2 ({file2})"  : file2,
            f"Count ({file2})"    : v2,
            "Delta"               : v2 - v1,
        })
    return pd.DataFrame(rows)


def build_summary_changes(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]

    categories = sorted([
        c for c in set(df1["slot_category"]) | set(df2["slot_category"])
        if c not in ("nan", "ltd", "")
    ])

    # Build UWI → key map for each file (from Present In file)
    key_map1 = {row["UWI"]: make_key(row) for _, row in df1.iterrows()}
    key_map2 = {row["UWI"]: make_key(row) for _, row in df2.iterrows()}

    rows = []
    for cat in categories:
        uwi1 = set(df1[df1["slot_category"] == cat]["UWI"].dropna().unique())
        uwi2 = set(df2[df2["slot_category"] == cat]["UWI"].dropna().unique())
        for u in sorted(uwi2 - uwi1):
            rows.append({"Category": cat.upper(), "Status": "ADDED",
                         "UWI": u, "Key": key_map2.get(u, ""),
                         "Present In": file2, "Absent From": file1})
        for u in sorted(uwi1 - uwi2):
            rows.append({"Category": cat.upper(), "Status": "DROPPED",
                         "UWI": u, "Key": key_map1.get(u, ""),
                         "Present In": file1, "Absent From": file2})

    if not rows:
        return pd.DataFrame(columns=["Category", "Status", "UWI", "Key", "Present In", "Absent From"])
    return pd.DataFrame(rows)


def build_category_sheet(df1, df2, cat, file1, file2):
    def get_subset(df):
        sub = df[df["slot_category"] == cat].copy()
        sub["Slot_ID"] = sub["Asset_Name"].astype(str) + "_" + sub["UWI"].astype(str)
        sub = sub[["UWI", "Development_Area_Name", "Slot_ID", "Well_Name"]].drop_duplicates(subset=["UWI"])
        return sub.reset_index(drop=True)

    s1   = get_subset(df1)
    s2   = get_subset(df2)
    set1 = set(s1["UWI"])
    set2 = set(s2["UWI"])

    n = max(len(s1), len(s2), 1)

    def pad(lst, length):
        return lst + [""] * (length - len(lst))

    uwi1    = pad(s1["UWI"].tolist(), n)
    dev1    = pad(s1["Development_Area_Name"].tolist(), n)
    sid1    = pad(s1["Slot_ID"].tolist(), n)
    well1   = pad(s1["Well_Name"].tolist(), n)
    uwi2    = pad(s2["UWI"].tolist(), n)
    dev2    = pad(s2["Development_Area_Name"].tolist(), n)
    sid2    = pad(s2["Slot_ID"].tolist(), n)
    well2   = pad(s2["Well_Name"].tolist(), n)

    flag1 = [1 if v and v in set2 else (0 if v else "") for v in uwi1]
    flag2 = [1 if v and v in set1 else (0 if v else "") for v in uwi2]

    f1_col = [file1 if v else "" for v in uwi1]
    f2_col = [file2 if v else "" for v in uwi2]

    return pd.DataFrame({
        f"Excel Sheet 1 ({file1})"                        : f1_col,
        f"Development Area Name (Sheet1)"                 : dev1,
        f"Slot_ID (Sheet1)"                               : sid1,
        f"Well Name (Sheet1)"                             : well1,
        f"{cat}s in Excel Sheet 1"                        : uwi1,
        f"Flag 1 ({cat} in Sheet1 found in Sheet2)"       : flag1,
        f"Excel Sheet 2 ({file2})"                        : f2_col,
        f"Development Area Name (Sheet2)"                 : dev2,
        f"Slot_ID (Sheet2)"                               : sid2,
        f"Well Name (Sheet2)"                             : well2,
        f"{cat}s in Excel Sheet 2"                        : uwi2,
        f"Flag 2 ({cat} in Sheet2 found in Sheet1)"       : flag2,
    })


def build_ltd_sheet(df1, df2, file1, file2):
    def get_unique(df):
        sub = df[df["slot_category"] == "ltd"].copy()
        sub["LTD_Key"] = sub.apply(make_ltd_key, axis=1)
        return sub[["LTD_Key"]].drop_duplicates().reset_index(drop=True)

    s1   = get_unique(df1)
    s2   = get_unique(df2)
    set1 = set(s1["LTD_Key"])
    set2 = set(s2["LTD_Key"])

    n = max(len(s1), len(s2), 1)

    def pad(lst, length):
        return lst + [""] * (length - len(lst))

    keys1 = pad(s1["LTD_Key"].tolist(), n)
    keys2 = pad(s2["LTD_Key"].tolist(), n)
    flag1 = [1 if v and v in set2 else (0 if v else "") for v in keys1]
    flag2 = [1 if v and v in set1 else (0 if v else "") for v in keys2]

    # Fill file name only on rows that have a key value
    f1_col = [file1 if v else "" for v in keys1]
    f2_col = [file2 if v else "" for v in keys2]

    return pd.DataFrame({
        f"Excel Sheet 1 ({file1})"                   : f1_col,
        "Key (Asset_Area_Formation_Slot) Sheet1"     : keys1,
        "Flag 1 (Key in Sheet1 found in Sheet2)"     : flag1,
        f"Excel Sheet 2 ({file2})"                   : f2_col,
        "Key (Asset_Area_Formation_Slot) Sheet2"     : keys2,
        "Flag 2 (Key in Sheet2 found in Sheet1)"     : flag2,
    })


def build_asset_analysis(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """
    Per Asset Name (unique):
      - Net Acreage  : single value per Asset Name (first occurrence)
      - Royalty      : Royalty_Rate + Overriding_Royalty_Rate (first occurrence, same for all rows)
    Columns: File1 | Asset Name | Net Acreage | Royalty | File2 | Asset Name | Net Acreage | Royalty | Delta Net Acreage | Delta Royalty
    Only matching Asset Names get delta values.
    """
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]

    def summarise(df):
        grp = (df.groupby("Asset_Name", sort=True)
                 .first()
                 .reset_index()[["Asset_Name", "Net_Acreage", "Royalty_Rate", "Overriding_Royalty_Rate"]])
        grp["Royalty Rate + ORRI"] = pd.to_numeric(grp["Royalty_Rate"], errors="coerce").fillna(0) + \
                         pd.to_numeric(grp["Overriding_Royalty_Rate"], errors="coerce").fillna(0)
        grp["Net_Acreage"] = pd.to_numeric(grp["Net_Acreage"], errors="coerce")
        return grp[["Asset_Name", "Net_Acreage", "Royalty Rate + ORRI"]].set_index("Asset_Name")

    s1 = summarise(df1)
    s2 = summarise(df2)

    all_assets = sorted(set(s1.index) | set(s2.index))
    rows = []
    for asset in all_assets:
        in1 = asset in s1.index
        in2 = asset in s2.index

        na1  = s1.loc[asset, "Net_Acreage"]        if in1 else ""
        roy1 = s1.loc[asset, "Royalty Rate + ORRI"] if in1 else ""
        na2  = s2.loc[asset, "Net_Acreage"]        if in2 else ""
        roy2 = s2.loc[asset, "Royalty Rate + ORRI"] if in2 else ""

        # Delta only for matching assets
        if in1 and in2:
            delta_na  = round(float(na2)  - float(na1),  6) if na1  != "" and na2  != "" else ""
            delta_roy = round(float(roy2) - float(roy1), 6) if roy1 != "" and roy2 != "" else ""
        else:
            delta_na = delta_roy = ""

        rows.append({
            f"File 1 ({file1})"                      : file1 if in1 else "",
            f"Asset Name ({file1})"                  : asset if in1 else "",
            f"Net Acreage ({file1})"                 : na1,
            f"Royalty Rate + ORRI ({file1})"         : roy1,
            f"File 2 ({file2})"                      : file2 if in2 else "",
            f"Asset Name ({file2})"                  : asset if in2 else "",
            f"Net Acreage ({file2})"                 : na2,
            f"Royalty Rate + ORRI ({file2})"         : roy2,
            "Delta Net Acreage"                      : delta_na,
            "Delta Royalty Rate + ORRI"              : delta_roy,
        })

    return pd.DataFrame(rows)


def build_category_changes(df1, df2):
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]

    map1 = df1.groupby("UWI")["slot_category"].first().reset_index().rename(columns={"slot_category": "cat1"})
    map2 = df2.groupby("UWI")["slot_category"].first().reset_index().rename(columns={"slot_category": "cat2"})

    merged  = map1.merge(map2, on="UWI")
    changed = merged[merged["cat1"] != merged["cat2"]].copy()

    if changed.empty:
        return pd.DataFrame(columns=["UWI", "File 1", "Category in File 1",
                                     "File 2", "Category in File 2", "Change"])

    changed["File 1"]             = file1
    changed["File 2"]             = file2
    changed["Category in File 1"] = changed["cat1"].str.upper()
    changed["Category in File 2"] = changed["cat2"].str.upper()
    changed["Change"]             = changed["Category in File 1"] + " → " + changed["Category in File 2"]
    return changed[["UWI", "File 1", "Category in File 1",
                    "File 2", "Category in File 2", "Change"]].reset_index(drop=True)


def build_changes(df1, df2):
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]
    all_cats = sorted(set(df1["slot_category"]) | set(df2["slot_category"]) - {"nan", ""})

    changes = {"file1": file1, "file2": file2, "categories": []}
    for cat in all_cats:
        if cat == "nan":
            continue
        uwi1    = set(df1[df1["slot_category"] == cat]["UWI"].dropna().unique())
        uwi2    = set(df2[df2["slot_category"] == cat]["UWI"].dropna().unique())
        added   = uwi2 - uwi1
        removed = uwi1 - uwi2
        changes["categories"].append({
            "category": cat.upper(),
            "count1"  : len(uwi1),
            "count2"  : len(uwi2),
            "added"   : len(added),
            "removed" : len(removed),
            "delta"   : len(uwi2) - len(uwi1),
        })
    return changes


# ── Styling ───────────────────────────────────────────────────────────────────
def style_header(ws, n_cols, color="4472C4"):
    fill   = PatternFill("solid", fgColor=color)
    font   = Font(bold=True, color="FFFFFF")
    align  = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, align, border


def style_category_header(ws, n_cols):
    """Alternating dark/light blue headers to differentiate File1 vs File2 columns."""
    dark_fill  = PatternFill("solid", fgColor="1565C0")
    light_fill = PatternFill("solid", fgColor="90CAF9")
    dark_font  = Font(bold=True, color="FFFFFF")
    light_font = Font(bold=True, color="0D47A1")
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # First 5 cols = Sheet1 group (dark), last 5 = Sheet2 group (light)
    mid = n_cols // 2
    for c in range(1, n_cols + 1):
        cell       = ws.cell(row=1, column=c)
        is_sheet1  = c <= mid
        cell.fill  = dark_fill  if is_sheet1 else light_fill
        cell.font  = dark_font  if is_sheet1 else light_font
        cell.alignment = align
        cell.border    = border


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ── Write Output ──────────────────────────────────────────────────────────────
def write_output(summary, df1, df2, combined, raw1, raw2, output_path):
    file1 = df1["source_file"].iloc[0]
    file2 = df2["source_file"].iloc[0]

    categories  = sorted([c for c in combined["slot_category"].unique()
                           if c not in ("nan", "ltd", "")])
    cat_changes = build_category_changes(df1, df2)
    chg_df      = build_summary_changes(df1, df2)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Summary ──────────────────────────────────────────────────────────
        summary.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        style_header(ws, len(summary.columns))
        auto_width(ws)

        # Changes section
        summary_end_row = len(summary) + 2
        title_row       = summary_end_row + 1
        ws.cell(row=title_row, column=1, value="📋 UWI Changes Between Files")
        for c in range(1, len(chg_df.columns) + 1):
            cell = ws.cell(row=title_row, column=c)
            cell.font      = Font(bold=True, size=12, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="37474F")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        chg_header_row = title_row + 1
        for c, col_name in enumerate(chg_df.columns, start=1):
            cell            = ws.cell(row=chg_header_row, column=c, value=col_name)
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.fill       = PatternFill("solid", fgColor="546E7A")
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        green_fill = PatternFill("solid", fgColor="C8E6C9")
        red_fill   = PatternFill("solid", fgColor="FFCDD2")
        green_font = Font(bold=True, color="1B5E20")
        red_font   = Font(bold=True, color="B71C1C")

        for i, (_, row) in enumerate(chg_df.iterrows()):
            r        = chg_header_row + 1 + i
            is_added = row["Status"] == "ADDED"
            row_fill = green_fill if is_added else red_fill
            for c, val in enumerate(row, start=1):
                cell           = ws.cell(row=r, column=c, value=val)
                cell.fill      = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if c == 2:
                    cell.font = green_font if is_added else red_font
        auto_width(ws)

        # ── LTD sheet ────────────────────────────────────────────────────────
        ltd_df = build_ltd_sheet(df1, df2, file1, file2)
        ltd_df.to_excel(writer, sheet_name="ltd", index=False)
        ws = writer.sheets["ltd"]
        style_category_header(ws, len(ltd_df.columns))
        auto_width(ws)

        # ── Category sheets ───────────────────────────────────────────────────
        for cat in categories:
            cat_df = build_category_sheet(df1, df2, cat, file1, file2)
            label  = str(cat)[:31].replace("/", "-").replace("\\", "-")
            cat_df.to_excel(writer, sheet_name=label, index=False)
            ws = writer.sheets[label]
            style_category_header(ws, len(cat_df.columns))
            auto_width(ws)

        # ── Asset Analysis sheet ─────────────────────────────────────────────
        asset_df = build_asset_analysis(df1, df2)
        asset_df.to_excel(writer, sheet_name="Asset Analysis", index=False)
        ws = writer.sheets["Asset Analysis"]

        # Header: dark blue for File1 group, light blue for File2 group, green for deltas
        n_cols     = len(asset_df.columns)
        dark_fill  = PatternFill("solid", fgColor="1565C0")
        light_fill = PatternFill("solid", fgColor="90CAF9")
        green_fill = PatternFill("solid", fgColor="2E7D32")
        dark_font  = Font(bold=True, color="FFFFFF")
        light_font = Font(bold=True, color="0D47A1")
        green_font = Font(bold=True, color="FFFFFF")
        align      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin       = Side(style="thin")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            if c <= 4:
                cell.fill, cell.font = dark_fill,  dark_font
            elif c <= 8:
                cell.fill, cell.font = light_fill, light_font
            else:
                cell.fill, cell.font = green_fill, green_font
            cell.alignment, cell.border = align, border

        # Highlight delta cells — red if negative, green if positive
        pos_fill = PatternFill("solid", fgColor="C8E6C9")
        neg_fill = PatternFill("solid", fgColor="FFCDD2")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column in (9, 10) and cell.value != "" and cell.value is not None:
                    try:
                        cell.fill = pos_fill if float(cell.value) >= 0 else neg_fill
                    except (ValueError, TypeError):
                        pass

        auto_width(ws)

        # ── Category Changes sheet ────────────────────────────────────────────
        cat_changes.to_excel(writer, sheet_name="Category Changes", index=False)
        ws = writer.sheets["Category Changes"]
        style_header(ws, len(cat_changes.columns), color="E65100")
        auto_width(ws)
        orange_fill = PatternFill("solid", fgColor="FFE0B2")
        orange_font = Font(bold=True, color="E65100")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = orange_fill
            row[-1].font = orange_font

        # Disclaimer row at the bottom
        disclaimer_row = ws.max_row + 2
        disclaimer_text = (
            "⚠️  Disclaimer: This sheet only captures category changes between non-LTD categories "
            "(e.g. DUC → PDP, PERMIT → DUC). Conversions involving LTD (e.g. LTD → DUC, LTD → PERMIT, "
            "DUC → LTD) are not displayed here as LTD entries are tracked by slot key "
            "(Area + Asset + Formation + Slot Number) rather than by UWI."
        )
        cell = ws.cell(row=disclaimer_row, column=1, value=disclaimer_text)
        cell.font      = Font(bold=True, color="B71C1C", italic=True)
        cell.fill      = PatternFill("solid", fgColor="FFF9C4")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[disclaimer_row].height = 45
        ws.merge_cells(start_row=disclaimer_row, start_column=1,
                       end_row=disclaimer_row, end_column=len(cat_changes.columns))

        # ── Raw Slots Allocation sheets ───────────────────────────────────────
        raw1.to_excel(writer, sheet_name=f"Slots_Allocation_{file1}"[:31], index=False)
        ws = writer.sheets[f"Slots_Allocation_{file1}"[:31]]
        style_header(ws, len(raw1.columns), color="37474F")
        auto_width(ws)

        raw2.to_excel(writer, sheet_name=f"Slots_Allocation_{file2}"[:31], index=False)
        ws = writer.sheets[f"Slots_Allocation_{file2}"[:31]]
        style_header(ws, len(raw2.columns), color="37474F")
        auto_width(ws)


# ── Summary Popup ─────────────────────────────────────────────────────────────
class SummaryPopup(tk.Toplevel):
    def __init__(self, parent, changes, save_path):
        super().__init__(parent)
        self.title("✅ Comparison Complete — Summary of Changes")
        self.configure(bg="#F5F5F5")
        self.resizable(False, False)
        self._build(changes, save_path)
        self._center(parent)
        self.grab_set()

    def _center(self, parent):
        self.update_idletasks()
        w, h = 600, 480
        px = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        py = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{px}+{py}")

    def _build(self, changes, save_path):
        file1, file2 = changes["file1"], changes["file2"]
        tk.Label(self, text="✅ Comparison Complete",
                 font=("Helvetica", 14, "bold"), bg="#F5F5F5", fg="#1A237E").pack(pady=(16, 2))
        tk.Label(self, text=f"Comparing:  {file1}  vs  {file2}",
                 font=("Helvetica", 8), bg="#F5F5F5", fg="#555").pack()
        tk.Frame(self, height=1, bg="#BDBDBD").pack(fill="x", padx=20, pady=10)

        container = tk.Frame(self, bg="#F5F5F5"); container.pack(fill="both", expand=True, padx=20)
        canvas    = tk.Canvas(container, bg="#F5F5F5", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        inner     = tk.Frame(canvas, bg="#F5F5F5")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        headers    = ["Category", f"Count\n({file1[:18]})", f"Count\n({file2[:18]})",
                      "Delta", "Added\n(new in File2)", "Removed\n(not in File2)"]
        col_widths = [10, 12, 12, 8, 14, 14]
        for c, (h, w) in enumerate(zip(headers, col_widths)):
            tk.Label(inner, text=h, font=("Helvetica", 8, "bold"),
                     bg="#1A237E", fg="white", width=w, relief="flat",
                     padx=4, pady=6, justify="center").grid(row=0, column=c, padx=1, pady=1)

        for r, cat in enumerate(changes["categories"], start=1):
            row_bg = "#FFFFFF" if r % 2 == 0 else "#F3F4FF"
            delta  = cat["delta"]
            if delta > 0:   delta_txt, delta_fg = f"+{delta}", "#2E7D32"
            elif delta < 0: delta_txt, delta_fg = str(delta), "#C62828"
            else:           delta_txt, delta_fg = "0", "#555"
            vals = [cat["category"], cat["count1"], cat["count2"],
                    delta_txt, cat["added"], cat["removed"]]
            fgs  = ["#222", "#222", "#222", delta_fg, "#2E7D32", "#C62828"]
            for c, (val, fg, w) in enumerate(zip(vals, fgs, col_widths)):
                tk.Label(inner, text=str(val), font=("Helvetica", 9),
                         bg=row_bg, fg=fg, width=w,
                         padx=4, pady=5).grid(row=r, column=c, padx=1, pady=1)

        tk.Frame(self, height=1, bg="#BDBDBD").pack(fill="x", padx=20, pady=8)
        tk.Label(self, text=f"💾  Saved to: {save_path}",
                 font=("Helvetica", 8), bg="#F5F5F5", fg="#555", wraplength=560).pack()
        tk.Button(self, text="Close", command=self.destroy,
                  bg="#1565C0", fg="white", font=("Helvetica", 10, "bold"),
                  relief="flat", padx=20, pady=6).pack(pady=12)


# ── GUI ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Evaluation Slot Comparison Tool")
        self.resizable(False, False)
        self.configure(bg="#F5F5F5")
        self._center()
        self._build_ui()
        self.file1 = self.file2 = None

    def _center(self):
        self.update_idletasks()
        w, h = 520, 460
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        pad = dict(padx=20, pady=8)
        tk.Label(self, text="📊 Evaluation Slot Comparison Tool",
                 font=("Helvetica", 16, "bold"), bg="#F5F5F5", fg="#1A237E").pack(pady=(20, 4))
        tk.Label(self, text="Upload two Excel files with a 'Slots Allocations' sheet",
                 font=("Helvetica", 9), bg="#F5F5F5", fg="#555").pack()

        desc_frame = tk.Frame(self, bg="#E8EAF6", bd=1, relief="solid")
        desc_frame.pack(fill="x", padx=20, pady=(8, 0))
        desc_text = (
            "📋  What you'll get in the output file:\n"
            "  • Summary  —  UWI counts grouped by file & slot category, with LTD count\n"
            "  • LTD  —  Side-by-side key comparisons with match flags\n"
            "  • DUC / Permit / PDP…  —  Side-by-side UWI lists with area, asset & flags\n"
            "  • Category Changes  —  UWIs that moved between slot categories\n"
            "  • Raw Slots Allocation sheets for both files"
        )
        tk.Label(desc_frame, text=desc_text, font=("Helvetica", 8), bg="#E8EAF6",
                 fg="#1A237E", justify="left", anchor="w", padx=10, pady=8).pack(fill="x")

        f1 = tk.Frame(self, bg="#F5F5F5"); f1.pack(fill="x", **pad)
        tk.Label(f1, text="File 1:", width=7, anchor="w", bg="#F5F5F5").pack(side="left")
        self.lbl1 = tk.Label(f1, text="No file selected", fg="#999",
                             bg="white", relief="sunken", anchor="w", width=38)
        self.lbl1.pack(side="left", padx=(0, 8))
        tk.Button(f1, text="Browse", command=self._pick1,
                  bg="#1565C0", fg="white", relief="flat", padx=8).pack(side="left")

        f2 = tk.Frame(self, bg="#F5F5F5"); f2.pack(fill="x", **pad)
        tk.Label(f2, text="File 2:", width=7, anchor="w", bg="#F5F5F5").pack(side="left")
        self.lbl2 = tk.Label(f2, text="No file selected", fg="#999",
                             bg="white", relief="sunken", anchor="w", width=38)
        self.lbl2.pack(side="left", padx=(0, 8))
        tk.Button(f2, text="Browse", command=self._pick2,
                  bg="#1565C0", fg="white", relief="flat", padx=8).pack(side="left")

        self.progress = ttk.Progressbar(self, mode="indeterminate", length=460)
        self.progress.pack(pady=(12, 0))
        self.status = tk.Label(self, text="", bg="#F5F5F5", fg="#333", font=("Helvetica", 9))
        self.status.pack()
        self.btn_run = tk.Button(self, text="▶  Generate Comparison",
                                 command=self._run, state="disabled",
                                 bg="#2E7D32", fg="white", font=("Helvetica", 11, "bold"),
                                 relief="flat", padx=16, pady=8)
        self.btn_run.pack(pady=(10, 0))

    def _pick1(self):
        path = filedialog.askopenfilename(title="Select File 1",
                                          filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file1 = path
            self.lbl1.config(text=os.path.basename(path), fg="#222")
            self._check_ready()

    def _pick2(self):
        path = filedialog.askopenfilename(title="Select File 2",
                                          filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file2 = path
            self.lbl2.config(text=os.path.basename(path), fg="#222")
            self._check_ready()

    def _check_ready(self):
        if self.file1 and self.file2:
            self.btn_run.config(state="normal")

    def _run(self):
        self.btn_run.config(state="disabled")
        self.status.config(text="Processing …", fg="#1565C0")
        self.progress.start(10)
        threading.Thread(target=self._process, daemon=True).start()

    def _process(self):
        try:
            df1      = read_slots(self.file1)
            df2      = read_slots(self.file2)
            raw1     = read_raw(self.file1)
            raw2     = read_raw(self.file2)
            combined = pd.concat([df1, df2], ignore_index=True)
            summary  = build_summary(df1, df2)
            changes  = build_changes(df1, df2)

            save_path = filedialog.asksaveasfilename(
                title="Save comparison as …",
                initialfile=OUTPUT_FILE,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not save_path:
                self._done("Save cancelled.", "#E65100")
                return

            write_output(summary, df1, df2, combined, raw1, raw2, save_path)
            self._done(f"✅ Saved: {os.path.basename(save_path)}", "#2E7D32")
            self.after(0, lambda: SummaryPopup(self, changes, save_path))

        except Exception as e:
            self._done(f"❌ Error: {e}", "#C62828")
            messagebox.showerror("Error", str(e))

    def _done(self, msg, color):
        self.progress.stop()
        self.status.config(text=msg, fg=color)
        self.btn_run.config(state="normal")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()
